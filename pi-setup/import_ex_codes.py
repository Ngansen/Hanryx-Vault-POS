"""
import_ex_codes.py — EX-era online serial code loader (U4)

Reads BOTH `Pokemon TCG ex Serial Codes.xlsx` (English) and
`Pokemon TCG ex Serial Codes - Japanese.xlsx` (Japanese) from your
Card-Database repo, and loads them into `src_eng_ex_codes` and
`src_jp_ex_codes` respectively.

Each xlsx has one sheet per set (Delta Species, Legend Maker, Holon
Phantoms, Crystal Guardians, Dragon Frontiers, Power Keepers, Nintendo
Promos, Jumbo Cards, ...). Total ~6,900 cards per language.

Schema per sheet (English):
  Name, Type, HP, Stage, #, Rarity, Code 1, Code 2 (sometimes),
  Code 3 (sometimes), RH Code

Why this matters: the EX-era cards (2003–2007) had unique online
codes for the now-defunct Pokémon TCG online client. Customers still
ask whether their code is unredeemed — and to verify it's a real
card vs a fake / re-print, we need to look it up. These tables also
let the consolidator detect the "Reverse Holo" variant (RH Code).
"""
from __future__ import annotations

import argparse
import json
import logging
import os
import sys
import tempfile
import time
import unicodedata
from pathlib import Path

import psycopg2
import psycopg2.extras

from unified.schema import init_unified_schema
from unified.sources import fetch_source

log = logging.getLogger("import_ex_codes")

# Header → logical-name. Identical between English and JP versions
# except the JP file has an extra "Japanese Name" / "JP Name" column
# in some sheets.
COLUMN_MAP: dict[str, list[str]] = {
    "card_name":     ["name", "card name", "english name", "name en"],
    "card_name_jp":  ["japanese name", "jp name", "japanese", "name jp", "name ja"],
    "card_type":     ["type", "card type"],
    "hp":            ["hp"],
    "stage":         ["stage"],
    "card_number":   ["#", "no", "no.", "number", "card number"],
    "rarity":        ["rarity"],
    "code_1":        ["code 1", "code1", "code"],
    "code_2":        ["code 2", "code2"],
    "code_3":        ["code 3", "code3"],
    "rh_code":       ["rh code", "rh", "reverse holo code", "reverse holo"],
}


def _norm(h) -> str:
    if h is None:
        return ""
    s = str(h).strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return s


def _build_index(headers: list) -> dict[str, int]:
    norm_headers = [_norm(h) for h in headers]
    out: dict[str, int] = {}
    for logical, cands in COLUMN_MAP.items():
        for cand in cands:
            cn = _norm(cand)
            for i, h in enumerate(norm_headers):
                if h == cn:
                    out[logical] = i
                    break
            if logical in out:
                break
    return out


def _safe_int(v):
    if v is None or v == "":
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def _safe_str(v) -> str:
    return "" if v is None else str(v).strip()


def _ingest(db_conn, xlsx_path: Path, *, table: str, has_jp_name: bool,
            source_label: str) -> dict:
    """Bulk-load every sheet in xlsx into `table`."""
    import openpyxl
    cur = db_conn.cursor()
    cur.execute(f"DELETE FROM {table} WHERE source_file = %s", (source_label,))

    now = int(time.time())
    inserted = 0
    skipped_sheets: list[str] = []

    # Pre-build the INSERT SQL based on table shape.
    if table == "src_eng_ex_codes":
        insert_sql = f"""
            INSERT INTO {table}
              (source_file, set_name, card_name, card_type, hp, stage,
               card_number, rarity, code_1, code_2, code_3, rh_code,
               raw_row, imported_at)
            VALUES %s
        """
    elif table == "src_jp_ex_codes":
        insert_sql = f"""
            INSERT INTO {table}
              (source_file, set_name, card_name_jp, card_name_en,
               card_type, hp, stage, card_number, rarity,
               code_1, code_2, code_3, rh_code, raw_row, imported_at)
            VALUES %s
        """
    else:
        raise ValueError(f"unsupported table: {table}")

    batch: list[tuple] = []
    BATCH_SIZE = 500

    def flush():
        nonlocal inserted
        if not batch:
            return
        psycopg2.extras.execute_values(cur, insert_sql, batch, page_size=500)
        inserted += len(batch)
        batch.clear()

    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            header = None
            idx: dict[str, int] = {}
            for row in ws.iter_rows(values_only=True):
                if row is None:
                    continue
                if header is None:
                    if all(c in (None, "") for c in row):
                        continue
                    header = list(row)
                    idx = _build_index(header)
                    if "card_name" not in idx and "card_name_jp" not in idx:
                        skipped_sheets.append(sheet_name)
                        break
                    continue
                if all(c in (None, "") for c in row):
                    continue

                def field(name: str):
                    i = idx.get(name)
                    if i is None or i >= len(row):
                        return None
                    return row[i]

                raw = {str(header[i] or f"c{i}").strip(): row[i]
                       for i in range(min(len(header), len(row)))}

                if table == "src_eng_ex_codes":
                    batch.append((
                        source_label,
                        sheet_name,
                        _safe_str(field("card_name")),
                        _safe_str(field("card_type")),
                        _safe_int(field("hp")),
                        _safe_str(field("stage")),
                        _safe_str(field("card_number")),
                        _safe_str(field("rarity")),
                        _safe_str(field("code_1")),
                        _safe_str(field("code_2")),
                        _safe_str(field("code_3")),
                        _safe_str(field("rh_code")),
                        json.dumps(raw, ensure_ascii=False, default=str),
                        now,
                    ))
                else:  # src_jp_ex_codes
                    batch.append((
                        source_label,
                        sheet_name,
                        _safe_str(field("card_name_jp")) or _safe_str(field("card_name")),
                        _safe_str(field("card_name")) if has_jp_name else "",
                        _safe_str(field("card_type")),
                        _safe_int(field("hp")),
                        _safe_str(field("stage")),
                        _safe_str(field("card_number")),
                        _safe_str(field("rarity")),
                        _safe_str(field("code_1")),
                        _safe_str(field("code_2")),
                        _safe_str(field("code_3")),
                        _safe_str(field("rh_code")),
                        json.dumps(raw, ensure_ascii=False, default=str),
                        now,
                    ))
                if len(batch) >= BATCH_SIZE:
                    flush()
    finally:
        wb.close()
    flush()
    db_conn.commit()
    return {"inserted": inserted, "skipped_sheets": skipped_sheets}


def import_ex_codes(db_conn, *, force: bool = False) -> dict:
    init_unified_schema(db_conn)

    cur = db_conn.cursor()
    cur.execute("SELECT COUNT(*) FROM src_eng_ex_codes")
    eng_pre = int(cur.fetchone()[0])
    cur.execute("SELECT COUNT(*) FROM src_jp_ex_codes")
    jp_pre = int(cur.fetchone()[0])

    if (eng_pre > 0 or jp_pre > 0) and not force:
        log.info("[ex-codes] already populated (eng=%d jp=%d) — skipping", eng_pre, jp_pre)
        return {"skipped": True, "eng_pre": eng_pre, "jp_pre": jp_pre}

    workdir = Path(tempfile.mkdtemp(prefix="ex-codes-"))
    out: dict = {}
    try:
        eng_path = fetch_source("english_ex_codes", workdir / "eng.xlsx")
        out["english"] = _ingest(db_conn, eng_path, table="src_eng_ex_codes",
                                 has_jp_name=False,
                                 source_label="Pokemon TCG ex Serial Codes.xlsx")
        try:
            jp_path = fetch_source("japanese_ex_codes", workdir / "jp.xlsx")
            out["japanese"] = _ingest(db_conn, jp_path, table="src_jp_ex_codes",
                                       has_jp_name=True,
                                       source_label="Pokemon TCG ex Serial Codes - Japanese.xlsx")
        except Exception as e:
            log.warning("[ex-codes] japanese file fetch failed: %s", e)
            out["japanese"] = {"error": str(e)}
    finally:
        for n in ("eng.xlsx", "jp.xlsx"):
            f = workdir / n
            if f.exists():
                try: f.unlink()
                except Exception: pass
        try: workdir.rmdir()
        except Exception: pass

    return out


def main() -> int:
    logging.basicConfig(level=logging.INFO,
                        format="%(asctime)s %(levelname)s %(name)s: %(message)s")
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--force", action="store_true")
    args = ap.parse_args()

    url = os.environ.get("DATABASE_URL")
    if not url:
        print("DATABASE_URL is not set", file=sys.stderr)
        return 1
    with psycopg2.connect(url) as conn:
        print(json.dumps(import_ex_codes(conn, force=args.force), indent=2,
                         ensure_ascii=False, default=str))
    return 0


if __name__ == "__main__":
    sys.exit(main())
