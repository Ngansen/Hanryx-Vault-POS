"""
import_eng_xlsx.py — English `ALL English Pokémon Cards.xlsx` loader (U3)

Reads the ~32K-row master English catalogue from your Card-Database
repo into `src_eng_xlsx`. The xlsx has these columns (header on row 1):

    ID, Set, Number, PokéDex ID, Card Name, Type,
    Rarity / Variant, Other Pokémon in Artwork, EX Serial Number(s)

The importer is column-name driven (case-insensitive, accent-insensitive
matching) so a future column rename in the upstream xlsx doesn't break
the load — it just leaves the corresponding field empty.

Storage strategy:
  * Always store the full raw row in `raw_row` JSONB so the consolidator
    can reach back for any column we forgot about.
  * Promote the columns the consolidator currently uses into typed
    columns for fast indexed lookup.

Idempotency: a fresh run wipes src_eng_xlsx in a transaction, then
bulk-inserts. Cheaper than per-row UPSERT for 32K rows and avoids
duplicate ID issues if the upstream sheet adds rows mid-table.
"""
from __future__ import annotations

import argparse
import json
import logging
import os
import sys
import tempfile
import time
import unicodedata
from pathlib import Path

import psycopg2
import psycopg2.extras

from unified.schema import init_unified_schema
from unified.sources import fetch_source

log = logging.getLogger("import_eng_xlsx")

# Logical name → list of accepted xlsx column headers (case-insensitive,
# accent-stripped). First match wins.
COLUMN_MAP: dict[str, list[str]] = {
    "row_id":            ["id", "card id", "card_id"],
    "set_name":          ["set", "set name", "set_name"],
    "card_number":       ["number", "card number", "card_number", "no", "no."],
    "pokedex_id":        ["pokedex id", "pokedex_id", "pokedex no", "national dex"],
    "card_name":         ["card name", "name", "cardname"],
    "card_type":         ["type", "card type"],
    "rarity_variant":    ["rarity / variant", "rarity/variant", "rarity_variant", "rarity"],
    "other_pokemon":     ["other pokemon in artwork", "other pokémon in artwork",
                          "other_pokemon", "other pokemon"],
    "ex_serial_numbers": ["ex serial number(s)", "ex serial numbers", "ex serial number",
                          "ex_serial_numbers", "serial number"],
}


def _norm(h) -> str:
    """Normalise a header for fuzzy matching."""
    if h is None:
        return ""
    s = str(h).strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.replace("/", " / ").replace("  ", " ").strip()
    return s


def _build_index(headers: list) -> dict[str, int]:
    """Map logical-field → column-index, given a header row from the xlsx."""
    norm_headers = [_norm(h) for h in headers]
    out: dict[str, int] = {}
    for logical, candidates in COLUMN_MAP.items():
        for cand in candidates:
            cn = _norm(cand)
            for i, h in enumerate(norm_headers):
                if h == cn:
                    out[logical] = i
                    break
            if logical in out:
                break
    return out


def _safe_int(v):
    if v is None or v == "":
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def _safe_str(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _iter_rows(xlsx_path: Path):
    """Yield (sheet_name, header_index, raw_dict, row_tuple) for every
    data row across every sheet in the workbook."""
    import openpyxl
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            header = None
            idx: dict[str, int] = {}
            for row in ws.iter_rows(values_only=True):
                if row is None:
                    continue
                if header is None:
                    # First non-empty row is the header row.
                    if all(c in (None, "") for c in row):
                        continue
                    header = list(row)
                    idx = _build_index(header)
                    if not idx:
                        # No recognised columns — skip whole sheet.
                        log.warning("[eng-xlsx] sheet %r has no recognised columns; skipping",
                                    sheet_name)
                        break
                    continue
                # Data row.
                if all(c in (None, "") for c in row):
                    continue
                raw = {}
                for i, h in enumerate(header):
                    if i < len(row):
                        raw[str(h or f"col{i}").strip()] = row[i]
                yield sheet_name, idx, raw, row
    finally:
        wb.close()


def _ingest(db_conn, xlsx_path: Path, source_label: str) -> dict:
    """Truncate + bulk-insert src_eng_xlsx from xlsx_path. Returns counts."""
    cur = db_conn.cursor()
    cur.execute("BEGIN")
    cur.execute("DELETE FROM src_eng_xlsx WHERE source_file = %s", (source_label,))

    now = int(time.time())
    inserted = 0
    batch: list[tuple] = []
    BATCH_SIZE = 500

    insert_sql = """
        INSERT INTO src_eng_xlsx
          (source_file, sheet_name, row_id, set_name, card_number,
           pokedex_id, card_name, card_type, rarity_variant,
           other_pokemon, ex_serial_numbers, raw_row, imported_at)
        VALUES %s
    """

    def flush():
        nonlocal inserted
        if not batch:
            return
        psycopg2.extras.execute_values(cur, insert_sql, batch, template=None, page_size=500)
        inserted += len(batch)
        batch.clear()

    for sheet_name, idx, raw, row in _iter_rows(xlsx_path):
        def field(name: str):
            i = idx.get(name)
            if i is None or i >= len(row):
                return None
            return row[i]

        batch.append((
            source_label,
            sheet_name,
            _safe_str(field("row_id")),
            _safe_str(field("set_name")),
            _safe_str(field("card_number")),
            _safe_int(field("pokedex_id")),
            _safe_str(field("card_name")),
            _safe_str(field("card_type")),
            _safe_str(field("rarity_variant")),
            _safe_str(field("other_pokemon")),
            _safe_str(field("ex_serial_numbers")),
            json.dumps(raw, ensure_ascii=False, default=str),
            now,
        ))
        if len(batch) >= BATCH_SIZE:
            flush()
    flush()

    db_conn.commit()
    return {"inserted": inserted, "source_file": source_label}


def import_english_xlsx(db_conn, *, force: bool = False) -> dict:
    """Load `ALL English Pokémon Cards.xlsx` into src_eng_xlsx."""
    init_unified_schema(db_conn)

    cur = db_conn.cursor()
    cur.execute("SELECT COUNT(*) FROM src_eng_xlsx")
    pre_count = int(cur.fetchone()[0])
    if pre_count > 0 and not force:
        log.info("[eng-xlsx] src_eng_xlsx has %d rows — skipping (force=False)", pre_count)
        return {"skipped": True, "pre_count": pre_count}

    workdir = Path(tempfile.mkdtemp(prefix="eng-xlsx-"))
    try:
        path = fetch_source("english_all_cards", workdir / "english.xlsx")
        result = _ingest(db_conn, path, source_label="ALL English Pokémon Cards.xlsx")
    finally:
        f = workdir / "english.xlsx"
        if f.exists():
            try: f.unlink()
            except Exception: pass
        try: workdir.rmdir()
        except Exception: pass

    return result


def main() -> int:
    logging.basicConfig(level=logging.INFO,
                        format="%(asctime)s %(levelname)s %(name)s: %(message)s")
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--force", action="store_true")
    args = ap.parse_args()

    url = os.environ.get("DATABASE_URL")
    if not url:
        print("DATABASE_URL is not set", file=sys.stderr)
        return 1

    with psycopg2.connect(url) as conn:
        result = import_english_xlsx(conn, force=args.force)
    print(json.dumps(result, indent=2, ensure_ascii=False, default=str))
    return 0


if __name__ == "__main__":
    sys.exit(main())
