"""
import_jp_xlsx.py — Japanese spreadsheets loader (U5)

Reads `Japanese Pokemon Card Spreadsheet 2.0 1996-Dec 2017.xlsx`
(13,538 rows, the newer & more complete file) into `src_jp_xlsx`.

We deliberately load only the v2.0 file by default — the 1996-2016
file (11,605 rows) is a strict subset of v2.0 plus the first half of
2016 was re-released with corrections. Pass --include-v1 to load
both with `source_file` discriminator.

Schema per sheet (MASTER):
  Card, Era, Type, Rarity, Special Rarity, Release Date,
  Set Name ENG, Set Name JPN, Set #, Promotional Card #,
  Got it?, Personal Card Notes
"""
from __future__ import annotations

import argparse
import json
import logging
import os
import sys
import tempfile
import time
import unicodedata
from pathlib import Path

import psycopg2
import psycopg2.extras

from unified.schema import init_unified_schema
from unified.sources import fetch_source

log = logging.getLogger("import_jp_xlsx")

COLUMN_MAP: dict[str, list[str]] = {
    "card_name":      ["card", "card name", "name"],
    "era":            ["era"],
    "card_type":      ["type", "card type"],
    "rarity":         ["rarity"],
    "special_rarity": ["special rarity", "special_rarity"],
    "release_date":   ["release date", "release_date", "released"],
    "set_name_eng":   ["set name eng", "set name english", "set name (eng)",
                      "set name", "set"],
    "set_name_jpn":   ["set name jpn", "set name japanese", "set name (jpn)",
                      "set name jp"],
    "set_number":     ["set #", "set number", "set no", "set num", "set#"],
    "promo_number":   ["promotional card #", "promo card #", "promo #",
                      "promotional card number", "promo number"],
}


def _norm(h) -> str:
    if h is None: return ""
    s = str(h).strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.replace("\xa0", " ")
    return s


def _build_index(headers: list) -> dict[str, int]:
    norm_headers = [_norm(h) for h in headers]
    out: dict[str, int] = {}
    for logical, cands in COLUMN_MAP.items():
        for cand in cands:
            cn = _norm(cand)
            for i, h in enumerate(norm_headers):
                if h == cn:
                    out[logical] = i
                    break
            if logical in out:
                break
    return out


def _safe_str(v) -> str:
    if v is None: return ""
    return str(v).strip().replace("\xa0", " ")


def _ingest(db_conn, xlsx_path: Path, source_label: str) -> dict:
    import openpyxl
    cur = db_conn.cursor()
    cur.execute("DELETE FROM src_jp_xlsx WHERE source_file = %s", (source_label,))

    now = int(time.time())
    inserted = 0
    batch: list[tuple] = []
    BATCH_SIZE = 500
    insert_sql = """
        INSERT INTO src_jp_xlsx
          (source_file, card_name, era, card_type, rarity, special_rarity,
           release_date, set_name_eng, set_name_jpn, set_number,
           promo_number, raw_row, imported_at)
        VALUES %s
    """

    def flush():
        nonlocal inserted
        if not batch: return
        psycopg2.extras.execute_values(cur, insert_sql, batch, page_size=500)
        inserted += len(batch)
        batch.clear()

    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            header = None
            idx: dict[str, int] = {}
            for row in ws.iter_rows(values_only=True):
                if row is None: continue
                if header is None:
                    if all(c in (None, "") for c in row): continue
                    header = list(row)
                    idx = _build_index(header)
                    if "card_name" not in idx:
                        log.warning("[jp-xlsx] sheet %r: no Card column, skipping", sheet_name)
                        break
                    continue
                if all(c in (None, "") for c in row): continue

                def field(name: str):
                    i = idx.get(name)
                    if i is None or i >= len(row): return None
                    return row[i]

                raw = {str(header[i] or f"c{i}").strip(): row[i]
                       for i in range(min(len(header), len(row)))}
                batch.append((
                    source_label,
                    _safe_str(field("card_name")),
                    _safe_str(field("era")),
                    _safe_str(field("card_type")),
                    _safe_str(field("rarity")),
                    _safe_str(field("special_rarity")),
                    _safe_str(field("release_date")),
                    _safe_str(field("set_name_eng")),
                    _safe_str(field("set_name_jpn")),
                    _safe_str(field("set_number")),
                    _safe_str(field("promo_number")),
                    json.dumps(raw, ensure_ascii=False, default=str),
                    now,
                ))
                if len(batch) >= BATCH_SIZE: flush()
    finally:
        wb.close()
    flush()
    db_conn.commit()
    return {"inserted": inserted, "source_file": source_label}


def import_jp_xlsx(db_conn, *, force: bool = False, include_v1: bool = False) -> dict:
    init_unified_schema(db_conn)

    cur = db_conn.cursor()
    cur.execute("SELECT COUNT(*) FROM src_jp_xlsx")
    pre = int(cur.fetchone()[0])
    if pre > 0 and not force:
        log.info("[jp-xlsx] src_jp_xlsx has %d rows — skipping", pre)
        return {"skipped": True, "pre_count": pre}

    workdir = Path(tempfile.mkdtemp(prefix="jp-xlsx-"))
    out: dict = {}
    try:
        v2_path = fetch_source("japanese_master_v2", workdir / "v2.xlsx")
        out["v2"] = _ingest(db_conn, v2_path,
                            source_label="Japanese Pokemon Card Spreadsheet 2.0 1996-Dec 2017.xlsx")
        if include_v1:
            v1_path = fetch_source("japanese_master_v1", workdir / "v1.xlsx")
            out["v1"] = _ingest(db_conn, v1_path,
                                source_label="Japanese Pokemon Card Master List 1996 - May 2016.xlsx")
    finally:
        for n in ("v1.xlsx", "v2.xlsx"):
            f = workdir / n
            if f.exists():
                try: f.unlink()
                except Exception: pass
        try: workdir.rmdir()
        except Exception: pass

    return out


def main() -> int:
    logging.basicConfig(level=logging.INFO,
                        format="%(asctime)s %(levelname)s %(name)s: %(message)s")
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--force", action="store_true")
    ap.add_argument("--include-v1", action="store_true",
                    help="Also load the older 1996-May 2016 file (subset of v2)")
    args = ap.parse_args()

    url = os.environ.get("DATABASE_URL")
    if not url:
        print("DATABASE_URL is not set", file=sys.stderr); return 1
    with psycopg2.connect(url) as conn:
        print(json.dumps(import_jp_xlsx(conn, force=args.force, include_v1=args.include_v1),
                         indent=2, ensure_ascii=False, default=str))
    return 0


if __name__ == "__main__":
    sys.exit(main())
