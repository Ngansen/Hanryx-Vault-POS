"""
import_ref_mappings.py — Korean + Chinese Master Database loader (U2)

Reads the two "Global Master Database" xlsx files from your
Card-Database repo and loads them into the `ref_set_mapping` and
`ref_variant_terms` tables.

These are TINY files (~7 KB, ~30 rows total combined) but
conceptually critical: they're the Rosetta Stone that lets the
consolidator match a Korean / Chinese set name back to the canonical
TCGdex set ID, and recognise variant terms like "마스터볼 미러" /
"大師球鏡面" / "Master Ball Holo" as the same logical thing.

Each Master DB xlsx has 3 sheets:
  1_Set_Registry        Era / Set ID / Korean Set / English Set / Release
                        (Chinese version: Region / Set ID / CN Name /
                         EN Name / Release)
  2_Master_Card_Mapping Set ID / No. / KR Name / EN Name / Rarity / Variant
                        (Chinese version adds Region column up-front)
  3_Variant_Logic       Variant Type / KR Term / Internal Code
                        (Chinese version: Region / CN Term / EN Term / Code)

The card_mapping sheet (sheet 2) is currently NOT loaded into ref_*
because it's a tiny per-card sample, not exhaustive — the consolidator
prefers the per-language card tables for that data. We expose it via
the raw_row JSONB on the set_mapping row so it's available if needed.
"""
from __future__ import annotations

import argparse
import json
import logging
import os
import sys
import tempfile
import time
from pathlib import Path

import psycopg2

from unified.schema import init_unified_schema
from unified.sources import fetch_source

log = logging.getLogger("import_ref_mappings")


def _open_workbook(path):
    import openpyxl
    return openpyxl.load_workbook(str(path), read_only=True, data_only=True)


def _rows(ws) -> list[list]:
    """Materialise a sheet to a list of plain Python lists."""
    out = []
    for row in ws.iter_rows(values_only=True):
        out.append(list(row))
    return out


# ─── Korean Master DB ─────────────────────────────────────────────────────

def _ingest_korean(db_conn, xlsx_path: Path, now: int) -> dict:
    wb = _open_workbook(xlsx_path)
    cur = db_conn.cursor()
    set_rows_inserted = 0
    variant_rows_inserted = 0

    if "1_Set_Registry" in wb.sheetnames:
        rows = _rows(wb["1_Set_Registry"])
        # header: Era / Set ID / Korean Set / English Set / Release
        for r in rows[1:]:
            if not r or all(c in (None, "") for c in r):
                continue
            era, set_id, kr_name, en_name, release = (
                (r + [None] * 5)[:5]
            )
            if not set_id:
                continue
            cur.execute(
                """
                INSERT INTO ref_set_mapping
                  (set_id, era, name_en, name_kr, release_year, region, raw, imported_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s::jsonb, %s)
                ON CONFLICT (set_id) DO UPDATE SET
                  era = EXCLUDED.era,
                  name_en = COALESCE(NULLIF(EXCLUDED.name_en, ''), ref_set_mapping.name_en),
                  name_kr = COALESCE(NULLIF(EXCLUDED.name_kr, ''), ref_set_mapping.name_kr),
                  release_year = COALESCE(NULLIF(EXCLUDED.release_year, ''), ref_set_mapping.release_year),
                  imported_at = EXCLUDED.imported_at
                """,
                (
                    str(set_id).strip(),
                    str(era or "").strip(),
                    str(en_name or "").strip(),
                    str(kr_name or "").strip(),
                    str(release or "").strip(),
                    "Korea",
                    json.dumps({"era": era, "kr": kr_name, "en": en_name, "release": release}, ensure_ascii=False, default=str),
                    now,
                ),
            )
            set_rows_inserted += 1

    if "3_Variant_Logic" in wb.sheetnames:
        rows = _rows(wb["3_Variant_Logic"])
        # header: Variant Type / KR Term / Internal Code
        for r in rows[1:]:
            if not r or all(c in (None, "") for c in r):
                continue
            en_term, kr_term, code = (r + [None] * 3)[:3]
            if not code:
                continue
            cur.execute(
                """
                INSERT INTO ref_variant_terms
                  (variant_code, en_term, kr_term, description, imported_at)
                VALUES (%s, %s, %s, %s, %s)
                ON CONFLICT (variant_code) DO UPDATE SET
                  en_term = COALESCE(NULLIF(EXCLUDED.en_term, ''), ref_variant_terms.en_term),
                  kr_term = COALESCE(NULLIF(EXCLUDED.kr_term, ''), ref_variant_terms.kr_term),
                  imported_at = EXCLUDED.imported_at
                """,
                (
                    str(code).strip().upper(),
                    str(en_term or "").strip(),
                    str(kr_term or "").strip(),
                    str(en_term or "").strip(),
                    now,
                ),
            )
            variant_rows_inserted += 1

    wb.close()
    db_conn.commit()
    return {"set_mappings": set_rows_inserted, "variants": variant_rows_inserted}


# ─── Chinese Master DB ────────────────────────────────────────────────────

def _ingest_chinese(db_conn, xlsx_path: Path, now: int) -> dict:
    wb = _open_workbook(xlsx_path)
    cur = db_conn.cursor()
    set_rows_inserted = 0
    variant_rows_inserted = 0

    if "1_Set_Registry" in wb.sheetnames:
        rows = _rows(wb["1_Set_Registry"])
        # header: Region / Set ID / CN Name / EN Name / Release
        for r in rows[1:]:
            if not r or all(c in (None, "") for c in r):
                continue
            region, set_id, cn_name, en_name, release = (r + [None] * 5)[:5]
            if not set_id:
                continue
            region_s = str(region or "").strip()
            # Determine which CN field to populate based on region
            chs = cn_name if region_s.lower().startswith("simplified") else None
            cht = cn_name if region_s.lower().startswith("traditional") else None
            cur.execute(
                """
                INSERT INTO ref_set_mapping
                  (set_id, name_en, name_chs, name_cht, release_year, region, raw, imported_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s::jsonb, %s)
                ON CONFLICT (set_id) DO UPDATE SET
                  name_en = COALESCE(NULLIF(EXCLUDED.name_en, ''), ref_set_mapping.name_en),
                  name_chs = COALESCE(NULLIF(EXCLUDED.name_chs, ''), ref_set_mapping.name_chs),
                  name_cht = COALESCE(NULLIF(EXCLUDED.name_cht, ''), ref_set_mapping.name_cht),
                  region = COALESCE(NULLIF(EXCLUDED.region, ''), ref_set_mapping.region),
                  release_year = COALESCE(NULLIF(EXCLUDED.release_year, ''), ref_set_mapping.release_year),
                  imported_at = EXCLUDED.imported_at
                """,
                (
                    str(set_id).strip(),
                    str(en_name or "").strip(),
                    str(chs or "").strip(),
                    str(cht or "").strip(),
                    str(release or "").strip(),
                    region_s,
                    json.dumps({"region": region_s, "cn": cn_name, "en": en_name, "release": release}, ensure_ascii=False, default=str),
                    now,
                ),
            )
            set_rows_inserted += 1

    if "3_Variant_Logic" in wb.sheetnames:
        rows = _rows(wb["3_Variant_Logic"])
        # header: Region / CN Term / EN Term / Code
        for r in rows[1:]:
            if not r or all(c in (None, "") for c in r):
                continue
            region, cn_term, en_term, code = (r + [None] * 4)[:4]
            if not code:
                continue
            region_s = str(region or "").strip()
            chs_t = cn_term if region_s.lower().startswith("simplified") else None
            cht_t = cn_term if region_s.lower().startswith("traditional") else None
            cur.execute(
                """
                INSERT INTO ref_variant_terms
                  (variant_code, en_term, chs_term, cht_term, description, imported_at)
                VALUES (%s, %s, %s, %s, %s, %s)
                ON CONFLICT (variant_code) DO UPDATE SET
                  en_term = COALESCE(NULLIF(EXCLUDED.en_term, ''), ref_variant_terms.en_term),
                  chs_term = COALESCE(NULLIF(EXCLUDED.chs_term, ''), ref_variant_terms.chs_term),
                  cht_term = COALESCE(NULLIF(EXCLUDED.cht_term, ''), ref_variant_terms.cht_term),
                  imported_at = EXCLUDED.imported_at
                """,
                (
                    str(code).strip().upper(),
                    str(en_term or "").strip(),
                    str(chs_t or "").strip(),
                    str(cht_t or "").strip(),
                    str(en_term or "").strip(),
                    now,
                ),
            )
            variant_rows_inserted += 1

    wb.close()
    db_conn.commit()
    return {"set_mappings": set_rows_inserted, "variants": variant_rows_inserted}


# ─── Public entry point ──────────────────────────────────────────────────

def import_ref_mappings(db_conn, *, force: bool = False) -> dict:
    """Load both Master DB files into ref_* tables.

    `force` is accepted for API parity with the other importers but
    ref_* upserts are always safe to re-run, so its only effect is
    bypassing the early-skip check.
    """
    init_unified_schema(db_conn)

    cur = db_conn.cursor()
    cur.execute("SELECT COUNT(*) FROM ref_set_mapping")
    pre_count = int(cur.fetchone()[0])
    if pre_count > 0 and not force:
        log.info("[ref-mappings] ref_set_mapping has %d rows — skipping (force=False)", pre_count)
        return {"skipped": True, "pre_count": pre_count}

    now = int(time.time())
    workdir = Path(tempfile.mkdtemp(prefix="ref-mappings-"))
    summary = {"korean": {}, "chinese": {}}
    try:
        kr_path = fetch_source("korean_master_db", workdir / "kr_master.xlsx")
        summary["korean"] = _ingest_korean(db_conn, kr_path, now)
        log.info("[ref-mappings] korean: %s", summary["korean"])

        cn_path = fetch_source("chinese_master_db", workdir / "cn_master.xlsx")
        summary["chinese"] = _ingest_chinese(db_conn, cn_path, now)
        log.info("[ref-mappings] chinese: %s", summary["chinese"])
    finally:
        # only clean up downloaded copies, not attached_assets paths
        for f in (workdir / "kr_master.xlsx", workdir / "cn_master.xlsx"):
            if f.exists():
                try: f.unlink()
                except Exception: pass
        try: workdir.rmdir()
        except Exception: pass

    cur.execute("SELECT COUNT(*) FROM ref_set_mapping")
    summary["total_set_mappings"] = int(cur.fetchone()[0])
    cur.execute("SELECT COUNT(*) FROM ref_variant_terms")
    summary["total_variants"] = int(cur.fetchone()[0])
    return summary


def main() -> int:
    logging.basicConfig(level=logging.INFO,
                        format="%(asctime)s %(levelname)s %(name)s: %(message)s")
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--force", action="store_true",
                    help="Re-import even if ref_set_mapping is non-empty")
    args = ap.parse_args()

    url = os.environ.get("DATABASE_URL")
    if not url:
        print("DATABASE_URL is not set", file=sys.stderr)
        return 1

    with psycopg2.connect(url) as conn:
        result = import_ref_mappings(conn, force=args.force)
    print(json.dumps(result, indent=2, ensure_ascii=False, default=str))
    return 0


if __name__ == "__main__":
    sys.exit(main())
